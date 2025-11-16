#!/usr/bin/env python3
import openpyxl
from datetime import datetime
import json
import sys

def extract_with_calculations(filepath, index_name):
    """Extrai dados calculando as fórmulas"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # Planilha BASE (mensal)
    ws_base = wb[f"{index_name}_BASE"]
    monthly_data = []
    
    for row in ws_base.iter_rows(min_row=2, max_row=ws_base.max_row-1):
        if row[0].value and isinstance(row[0].value, datetime):
            monthly_index = row[1].value if row[1].value else 0
            daily_index = row[4].value if row[4].value else 0
            monthly_data.append({
                "date": row[0].value.strftime("%Y-%m-01"),
                "monthly_index": float(monthly_index) if monthly_index else 0,
                "daily_index": float(daily_index) if daily_index else 0
            })
    
    # Planilha DIA (diária)
    ws_dia = wb[f"{index_name}_DIA"]
    daily_data = []
    
    for row in ws_dia.iter_rows(min_row=2, max_row=ws_dia.max_row):
        if row[0].value and isinstance(row[0].value, datetime):
            daily_index = row[2].value if row[2].value else 0
            accumulated = row[3].value if row[3].value else 0
            daily_data.append({
                "date": row[0].value.strftime("%Y-%m-%d"),
                "daily_index": float(daily_index) if daily_index else 0,
                "accumulated": float(accumulated) if accumulated else 0
            })
    
    return {
        "monthly": monthly_data,
        "daily": daily_data
    }

# Extrair dados
print("Extraindo dados INCC...")
incc_data = extract_with_calculations("/home/ubuntu/upload/INCC.xlsx", "INCC")
print(f"INCC: {len(incc_data['monthly'])} meses, {len(incc_data['daily'])} dias")

print("Extraindo dados IPCA...")
ipca_data = extract_with_calculations("/home/ubuntu/upload/IPCA.xlsx", "IPCA")
print(f"IPCA: {len(ipca_data['monthly'])} meses, {len(ipca_data['daily'])} dias")

# Salvar dados para uso posterior
with open("/home/ubuntu/escritorio-advocacia/scripts/indices_data.json", "w") as f:
    json.dump({
        "incc": incc_data,
        "ipca": ipca_data
    }, f)

print("Dados salvos em scripts/indices_data.json")
